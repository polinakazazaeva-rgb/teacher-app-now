import tkinter as tk
from tkinter import messagebox, ttk, scrolledtext, filedialog
import json
import os
from datetime import datetime
import re

DATA_FILE = "spravochnik_data.json"
STUDENTS_FILE = "students.json"

# ==================== УЧЕБНЫЙ КОНТЕНТ ====================
COURSE_CONTENT = """
ИНФОРМАЦИОННЫЕ ТЕХНОЛОГИИ — СОДЕРЖАНИЕ ПРОГРАММЫ (68 часов)

1. ОСНОВЫ ИНФОРМАЦИОННЫХ ТЕХНОЛОГИЙ (8 ч)
   – Понятие информации, виды информации (2 ч)
   – Измерение информации, единицы (2 ч)
   – Информационные процессы (2 ч)
   – История вычислительной техники (2 ч)

2. АРХИТЕКТУРА КОМПЬЮТЕРА (10 ч)
   – Устройство ПК, процессор (2 ч)
   – Память: виды и характеристики (2 ч)
   – Устройства ввода/вывода (2 ч)
   – Периферийное оборудование (2 ч)
   – Сборка и диагностика ПК (2 ч)

3. ПРОГРАММНОЕ ОБЕСПЕЧЕНИЕ (10 ч)
   – Системное ПО, ОС (2 ч)
   – Файловые системы (2 ч)
   – Прикладное ПО, офисные программы (2 ч)
   – Антивирусное ПО, защита (2 ч)
   – Установка и обновление ПО (2 ч)

4. КОМПЬЮТЕРНЫЕ СЕТИ (10 ч)
   – Локальные и глобальные сети (2 ч)
   – Топологии, протоколы (2 ч)
   – IP-адресация, DNS (2 ч)
   – Сетевые сервисы (2 ч)
   – Безопасность в сети (2 ч)

5. АЛГОРИТМИЗАЦИЯ (15 ч)
   – Понятие алгоритма (2 ч)
   – Линейные и ветвящиеся алгоритмы (3 ч)
   – Циклические алгоритмы (3 ч)
   – Языки программирования (2 ч)
   – Введение в Python (3 ч)
   – Контрольная работа (2 ч)

6. ИНФОРМАЦИОННАЯ БЕЗОПАСНОСТЬ (5 ч)
   – Основы ИБ (2 ч)
   – Шифрование (1 ч)
   – Вредоносное ПО (1 ч)
   – Правовые аспекты (1 ч)

7. ИТОГОВОЕ ПОВТОРЕНИЕ (5 ч)
   – Повторение (3 ч)
   – Итоговый зачёт (2 ч)
"""

LEARNING_OUTCOMES = """
ПЛАНИРУЕМЫЕ РЕЗУЛЬТАТЫ (Информационные технологии)

ЦЕЛИ ПРЕДМЕТА:
– Сформировать системное представление об информационных технологиях.
– Обучить практическим навыкам работы на ПК.
– Развить алгоритмическое и логическое мышление.
– Подготовить к использованию ИТ в профессиональной деятельности.

ЗАДАЧИ:
– Изучить устройство компьютера и периферию.
– Освоить офисные программы.
– Научиться работать в интернете.
– Изучить основы алгоритмизации и программирования.
– Познакомиться с основами информационной безопасности.

КОЛИЧЕСТВО ЧАСОВ: 68 (34 лекции + 30 практики + 4 контроля)

ВНЕСЁННЫЕ ИЗМЕНЕНИЯ:
– Добавлен раздел «Информационная безопасность» (5 ч).
– Увеличен объём практики по Python.
– Обновлено содержание раздела «Компьютерные сети».
"""

TEACHING_METHODS = """
ПОДРОБНЫЕ МЕТОДИКИ ОБУЧЕНИЯ

1. МЕТОД ПРОЕКТОВ
   Суть: студенты разрабатывают реальный проект от начала до защиты.
   Развивает самостоятельность, планирование, командную работу.
   Пример: создание веб-сайта, базы данных, мобильного приложения.

2. ПРОБЛЕМНОЕ ОБУЧЕНИЕ
   Суть: перед студентом ставится проблема, которую он решает.
   Стимулирует критическое мышление.
   Пример: «Как защитить компьютер от вирусов без антивируса?»

3. КЕЙС-МЕТОД
   Суть: разбор реальных ситуаций из практики.
   Учит анализировать и принимать решения.
   Пример: «Фирма потеряла данные. Что делать?»

4. ИГРОВЫЕ ТЕХНОЛОГИИ
   Суть: обучение через игру и соревнование.
   Пример: квиз по информатике, симулятор сборки ПК.

5. ПЕРЕВЁРНУТЫЙ КЛАСС
   Суть: теория дома (видео), практика в классе.

6. ДИФФЕРЕНЦИРОВАННОЕ ОБУЧЕНИЕ
   Суть: задания разного уровня сложности.
   Уровни: базовый (3), продвинутый (4), экспертный (5).
"""

# ==================== ФУНКЦИИ ====================
def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"users": []}
    return {"users": []}

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_students():
    if os.path.exists(STUDENTS_FILE):
        try:
            with open(STUDENTS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"students": []}
    return {"students": []}

def save_students(data):
    with open(STUDENTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def check_password_strength(password):
    if len(password) < 6:
        return False, "Слишком короткий (мин. 6 символов)", "#e74c3c"
    if not re.search(r'[A-ZА-Я]', password):
        return False, "Добавьте заглавную букву", "#e74c3c"
    if not re.search(r'\d', password):
        return False, "Добавьте цифру", "#e74c3c"
    return True, "Надёжный пароль", "#27ae60"

# ==================== ГЛАВНОЕ ПРИЛОЖЕНИЕ ====================
class TeacherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Справочник преподавателя — Информационные технологии")
        self.root.geometry("1200x700")
        self.root.configure(bg='#f0f7ff')
        
        self.data = load_data()
        self.current_user = None
        self.students_data = load_students()
        
        self.show_login()
    
    def clear(self):
        for widget in self.root.winfo_children():
            widget.destroy()
    
    # ==================== ВХОД ====================
    def show_login(self):
        self.clear()
        
        card = tk.Frame(self.root, bg='white', relief='flat')
        card.place(relx=0.5, rely=0.5, anchor='center', width=400, height=450)
        
        tk.Label(card, text="📚 СПРАВОЧНИК ПРЕПОДАВАТЕЛЯ", 
                font=("Segoe UI", 16, "bold"), bg='white', fg='#1a5276').pack(pady=30)
        tk.Label(card, text="Информационные технологии", 
                font=("Segoe UI", 10), bg='white', fg='#5d6d7e').pack()
        
        tk.Label(card, text="Логин", font=("Segoe UI", 11), bg='white', fg='#2c3e50').pack(anchor='w', padx=40, pady=(20,0))
        self.login_entry = tk.Entry(card, font=("Segoe UI", 11), bd=1, relief='solid')
        self.login_entry.pack(fill='x', padx=40, ipady=6)
        
        tk.Label(card, text="Пароль", font=("Segoe UI", 11), bg='white', fg='#2c3e50').pack(anchor='w', padx=40, pady=(10,0))
        self.password_entry = tk.Entry(card, font=("Segoe UI", 11), bd=1, relief='solid', show="*")
        self.password_entry.pack(fill='x', padx=40, ipady=6)
        
        tk.Button(card, text="ВОЙТИ", font=("Segoe UI", 11, "bold"), bg='#2c6e9e', fg='white',
                 command=self.do_login).pack(fill='x', padx=40, pady=20, ipady=8)
        tk.Button(card, text="СОЗДАТЬ АККАУНТ", font=("Segoe UI", 10), bg='white', fg='#2c6e9e',
                 command=self.show_register).pack()
    
    def show_register(self):
        self.clear()
        
        card = tk.Frame(self.root, bg='white', relief='flat')
        card.place(relx=0.5, rely=0.5, anchor='center', width=450, height=550)
        
        tk.Label(card, text="РЕГИСТРАЦИЯ", font=("Segoe UI", 16, "bold"),
                bg='white', fg='#1a5276').pack(pady=20)
        
        tk.Label(card, text="Логин", font=("Segoe UI", 10), bg='white', fg='#2c3e50').pack(anchor='w', padx=40, pady=(5,0))
        self.reg_login = tk.Entry(card, font=("Segoe UI", 10), bd=1, relief='solid')
        self.reg_login.pack(fill='x', padx=40, ipady=5)
        
        tk.Label(card, text="Ваше имя", font=("Segoe UI", 10), bg='white', fg='#2c3e50').pack(anchor='w', padx=40, pady=(10,0))
        self.reg_name = tk.Entry(card, font=("Segoe UI", 10), bd=1, relief='solid')
        self.reg_name.pack(fill='x', padx=40, ipady=5)
        
        tk.Label(card, text="Пароль", font=("Segoe UI", 10), bg='white', fg='#2c3e50').pack(anchor='w', padx=40, pady=(10,0))
        self.reg_pass = tk.Entry(card, font=("Segoe UI", 10), bd=1, relief='solid', show="*")
        self.reg_pass.pack(fill='x', padx=40, ipady=5)
        
        self.pass_status = tk.Label(card, text="", font=("Segoe UI", 9), bg='white')
        self.pass_status.pack(anchor='w', padx=40)
        self.reg_pass.bind('<KeyRelease>', self.update_strength)
        
        tk.Label(card, text="Подтвердите пароль", font=("Segoe UI", 10), bg='white', fg='#2c3e50').pack(anchor='w', padx=40, pady=(5,0))
        self.reg_confirm = tk.Entry(card, font=("Segoe UI", 10), bd=1, relief='solid', show="*")
        self.reg_confirm.pack(fill='x', padx=40, ipady=5)
        
        tk.Button(card, text="ЗАРЕГИСТРИРОВАТЬСЯ", font=("Segoe UI", 11, "bold"), bg='#27ae60', fg='white',
                 command=self.do_register).pack(fill='x', padx=40, pady=20, ipady=8)
        tk.Button(card, text="← НАЗАД", font=("Segoe UI", 9), bg='white', fg='#5d6d7e',
                 command=self.show_login).pack()
    
    def update_strength(self, event):
        pwd = self.reg_pass.get()
        if pwd:
            ok, msg, color = check_password_strength(pwd)
            self.pass_status.config(text=msg, fg=color)
        else:
            self.pass_status.config(text="")
    
    def do_register(self):
        login = self.reg_login.get().strip()
        name = self.reg_name.get().strip()
        pwd = self.reg_pass.get()
        confirm = self.reg_confirm.get()
        
        if not login or not name or not pwd:
            messagebox.showerror("Ошибка", "Заполните все поля!")
            return
        if pwd != confirm:
            messagebox.showerror("Ошибка", "Пароли не совпадают!")
            return
        ok, msg, _ = check_password_strength(pwd)
        if not ok:
            messagebox.showerror("Слабый пароль", msg)
            return
        for u in self.data["users"]:
            if u["login"] == login:
                messagebox.showerror("Ошибка", "Логин уже существует!")
                return
        
        self.data["users"].append({
            "login": login,
            "name": name,
            "password": pwd,
            "notes": ""
        })
        save_data(self.data)
        messagebox.showinfo("Успех", "Аккаунт создан!")
        self.show_login()
    
    def do_login(self):
        login = self.login_entry.get().strip()
        pwd = self.password_entry.get()
        for u in self.data["users"]:
            if u["login"] == login and u["password"] == pwd:
                self.current_user = u
                self.show_main()
                return
        messagebox.showerror("Ошибка", "Неверный логин или пароль!")
    
    # ==================== ГЛАВНОЕ МЕНЮ ====================
    def show_main(self):
        self.clear()
        
        # Верхняя панель
        header = tk.Frame(self.root, bg='#1a3c5c', height=80)
        header.pack(fill='x')
        tk.Label(header, text=f"👋 Здравствуйте, {self.current_user['name']}",
                font=("Segoe UI", 16, "bold"), bg='#1a3c5c', fg='white').place(x=20, y=25)
        tk.Button(header, text="🚪 ВЫХОД", bg='#c0392b', fg='white',
                 command=self.logout).place(x=1100, y=25)
        
        # Контейнер с карточками
        container = tk.Frame(self.root, bg='#f0f7ff')
        container.pack(fill='both', expand=True, padx=30, pady=30)
        
        # Карточки
        cards = [
            ("📖", "Содержание программы", self.show_content),
            ("🎯", "Планируемые результаты", self.show_outcomes),
            ("📋", "Методики обучения", self.show_methods),
            ("📅", "Расписание", self.show_schedule),
            ("📝", "Мои заметки", self.show_notes),
            ("📊", "Журнал успеваемости", self.show_journal),
        ]
        
        for i, (icon, title, command) in enumerate(cards):
            row, col = i // 3, i % 3
            card = tk.Frame(container, bg='white', relief='raised', bd=0, padx=20, pady=20)
            card.grid(row=row, column=col, padx=15, pady=15, sticky='nsew')
            
            tk.Label(card, text=title, font=("Segoe UI", 12, "bold"), bg='white', fg='#1a3c5c').pack()
            tk.Label(card, text=icon, font=("Segoe UI", 40), bg='white').pack(pady=5)
            tk.Button(card, text="Открыть", bg='#2c6e9e', fg='white',
                     command=command).pack(pady=10, ipadx=15, ipady=3)
        
        for i in range(3):
            container.columnconfigure(i, weight=1)
    
    # ==================== ПОКАЗ ТЕКСТА ====================
    def show_text(self, title, content):
        win = tk.Toplevel(self.root)
        win.title(title)
        win.geometry("800x600")
        win.configure(bg='white')
        
        tk.Label(win, text=title, font=("Segoe UI", 14, "bold"), bg='white', fg='#1a5276').pack(pady=10)
        
        text_area = scrolledtext.ScrolledText(win, font=("Segoe UI", 11), wrap='word', padx=15, pady=15)
        text_area.pack(fill='both', expand=True, padx=10, pady=10)
        text_area.insert('1.0', content)
        text_area.config(state='disabled')
        
        tk.Button(win, text="Закрыть", command=win.destroy).pack(pady=10)
    
    def show_content(self):
        self.show_text("Содержание программы", COURSE_CONTENT)
    
    def show_outcomes(self):
        self.show_text("Планируемые результаты", LEARNING_OUTCOMES)
    
    def show_methods(self):
        self.show_text("Методики обучения", TEACHING_METHODS)
    
    def show_schedule(self):
        win = tk.Toplevel(self.root)
        win.title("Расписание занятий")
        win.geometry("700x500")
        win.configure(bg='white')
        
        tk.Label(win, text="РАСПИСАНИЕ ЗАНЯТИЙ", font=("Segoe UI", 14, "bold"), 
                bg='white', fg='#1a5276').pack(pady=10)
        
        schedule_text = """
ПОНЕДЕЛЬНИК:
   09:00-10:30 - Информационные технологии (лекция)
   10:40-12:10 - Практикум по ПК (практика)

ВТОРНИК:
   09:00-10:30 - Программирование (лекция)
   10:40-12:10 - Лабораторные работы

СРЕДА:
   09:00-10:30 - Компьютерные сети (лекция)
   10:40-12:10 - Информационная безопасность

ЧЕТВЕРГ:
   09:00-10:30 - Базы данных (лекция)
   10:40-12:10 - Веб-технологии

ПЯТНИЦА:
   09:00-10:30 - Проектная деятельность
   10:40-12:10 - Консультации
"""
        text_area = scrolledtext.ScrolledText(win, font=("Segoe UI", 11), wrap='word', padx=15, pady=15)
        text_area.pack(fill='both', expand=True, padx=10, pady=10)
        text_area.insert('1.0', schedule_text)
        text_area.config(state='disabled')
        
        tk.Label(win, text="💡 Для синхронизации с Excel выберите файл:", 
                font=("Segoe UI", 9), bg='white', fg='#7f8c8d').pack()
        
        def load_excel():
            from tkinter import filedialog
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, data_only=True)
                    sheet = wb.active
                    lines = []
                    for row in sheet.iter_rows(values_only=True):
                        if row[0] and row[1]:
                            lines.append(f"{row[0]} | {row[1]}")
                    if lines:
                        text_area.config(state='normal')
                        text_area.delete('1.0', 'end')
                        text_area.insert('1.0', "\n".join(lines))
                        text_area.config(state='disabled')
                    else:
                        messagebox.showinfo("Инфо", "Нет данных в первых двух столбцах")
                except ImportError:
                    messagebox.showerror("Ошибка", "Установите openpyxl: pip install openpyxl")
                except Exception as e:
                    messagebox.showerror("Ошибка", str(e))
        
        tk.Button(win, text="📎 Загрузить из Excel", bg='#2c6e9e', fg='white',
                 command=load_excel).pack(pady=10)
        tk.Button(win, text="Закрыть", command=win.destroy).pack(pady=5)
    
    def show_notes(self):
        win = tk.Toplevel(self.root)
        win.title("Мои заметки")
        win.geometry("700x500")
        win.configure(bg='white')
        
        tk.Label(win, text="МОИ ЗАМЕТКИ", font=("Segoe UI", 14, "bold"), 
                bg='white', fg='#1a5276').pack(pady=10)
        
        text_area = scrolledtext.ScrolledText(win, font=("Segoe UI", 11), wrap='word', padx=15, pady=15)
        text_area.pack(fill='both', expand=True, padx=10, pady=10)
        
        if "notes" in self.current_user:
            text_area.insert('1.0', self.current_user["notes"])
        
        def save():
            self.current_user["notes"] = text_area.get('1.0', 'end-1c')
            for i, u in enumerate(self.data["users"]):
                if u["login"] == self.current_user["login"]:
                    self.data["users"][i] = self.current_user
                    break
            save_data(self.data)
            messagebox.showinfo("Успех", "Заметки сохранены!")
            win.destroy()
        
        tk.Button(win, text="💾 СОХРАНИТЬ", bg='#27ae60', fg='white', command=save).pack(pady=10)
        tk.Button(win, text="Закрыть", command=win.destroy).pack(pady=5)
    
    # ==================== ЖУРНАЛ УСПЕВАЕМОСТИ ====================
    def show_journal(self):
        win = tk.Toplevel(self.root)
        win.title("Журнал успеваемости")
        win.geometry("1000x700")
        win.configure(bg='#f0f7ff')
        
        nb = ttk.Notebook(win)
        nb.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Вкладка "Ученики"
        tab1 = tk.Frame(nb, bg='#f0f7ff')
        nb.add(tab1, text="📋 Ученики")
        
        left = tk.Frame(tab1, bg='white')
        left.pack(side='left', fill='both', expand=True, padx=5, pady=5)
        tk.Label(left, text="Список учеников", font=("Segoe UI", 12, "bold"), bg='white').pack()
        self.students_list = tk.Listbox(left, font=("Segoe UI", 10), height=15)
        self.students_list.pack(fill='both', expand=True, padx=10, pady=5)
        
        right = tk.Frame(tab1, bg='white', padx=10, pady=10)
        right.pack(side='right', fill='y')
        tk.Label(right, text="ФИО ученика", bg='white').pack(anchor='w')
        self.new_student_entry = tk.Entry(right, width=25)
        self.new_student_entry.pack(pady=5)
        tk.Button(right, text="➕ Добавить", bg='#27ae60', fg='white', 
                 command=self.add_student).pack(pady=2, fill='x')
        tk.Button(right, text="🗑 Удалить", bg='#e67e22', fg='white', 
                 command=self.remove_student).pack(pady=2, fill='x')
        
        # Вкладка "Оценки"
        tab2 = tk.Frame(nb, bg='#f0f7ff')
        nb.add(tab2, text="⭐ Оценки")
        
        top = tk.Frame(tab2, bg='white')
        top.pack(fill='x', padx=10, pady=10)
        tk.Label(top, text="Ученик:", bg='white').pack(side='left', padx=5)
        self.grade_student = ttk.Combobox(top, width=20)
        self.grade_student.pack(side='left', padx=5)
        tk.Label(top, text="Тема:", bg='white').pack(side='left', padx=5)
        self.grade_topic = tk.Entry(top, width=20)
        self.grade_topic.pack(side='left', padx=5)
        tk.Label(top, text="Оценка:", bg='white').pack(side='left', padx=5)
        self.grade_value = tk.Entry(top, width=5)
        self.grade_value.pack(side='left', padx=5)
        tk.Button(top, text="Поставить", bg='#2c6e9e', fg='white', 
                 command=self.add_grade).pack(side='left', padx=10)
        
        self.grades_tree = ttk.Treeview(tab2, columns=("Ученик", "Тема", "Оценка"), show='headings')
        self.grades_tree.heading("Ученик", text="Ученик")
        self.grades_tree.heading("Тема", text="Тема")
        self.grades_tree.heading("Оценка", text="Оценка")
        self.grades_tree.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Вкладка "Должники"
        tab3 = tk.Frame(nb, bg='#f0f7ff')
        nb.add(tab3, text="⚠️ Должники")
        
        self.debtors_tree = ttk.Treeview(tab3, columns=("Ученик", "Долги"), show='headings')
        self.debtors_tree.heading("Ученик", text="Ученик")
        self.debtors_tree.heading("Долги", text="Несданные темы")
        self.debtors_tree.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Вкладка "Статистика"
        tab4 = tk.Frame(nb, bg='#f0f7ff')
        nb.add(tab4, text="📊 Средний балл")
        
        self.stats_text = scrolledtext.ScrolledText(tab4, font=("Segoe UI", 11), wrap='word', bg='white')
        self.stats_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.refresh_students_list()
        self.refresh_grades_tree()
        self.refresh_debtors()
        self.refresh_stats()
        
        tk.Button(win, text="Обновить всё", bg='#2c6e9e', fg='white',
                 command=lambda: [self.refresh_students_list(), self.refresh_grades_tree(), 
                                 self.refresh_debtors(), self.refresh_stats()]).pack(pady=10)
    
    def refresh_students_list(self):
        self.students_list.delete(0, 'end')
        student_names = [s["name"] for s in self.students_data["students"]]
        for name in student_names:
            self.students_list.insert('end', name)
        self.grade_student['values'] = student_names
    
    def add_student(self):
        name = self.new_student_entry.get().strip()
        if not name:
            return
        for s in self.students_data["students"]:
            if s["name"] == name:
                messagebox.showerror("Ошибка", "Ученик уже есть")
                return
        self.students_data["students"].append({"name": name, "grades": {}})
        save_students(self.students_data)
        self.refresh_students_list()
        self.new_student_entry.delete(0, 'end')
    
    def remove_student(self):
        sel = self.students_list.curselection()
        if not sel:
            return
        name = self.students_list.get(sel[0])
        self.students_data["students"] = [s for s in self.students_data["students"] if s["name"] != name]
        save_students(self.students_data)
        self.refresh_students_list()
        self.refresh_grades_tree()
        self.refresh_debtors()
        self.refresh_stats()
    
    def add_grade(self):
        student = self.grade_student.get()
        topic = self.grade_topic.get().strip()
        grade = self.grade_value.get().strip()
        if not student or not topic or not grade:
            return
        for s in self.students_data["students"]:
            if s["name"] == student:
                if "grades" not in s:
                    s["grades"] = {}
                s["grades"][topic] = grade
                save_students(self.students_data)
                self.refresh_grades_tree()
                self.refresh_debtors()
                self.refresh_stats()
                self.grade_topic.delete(0, 'end')
                self.grade_value.delete(0, 'end')
                break
    
    def refresh_grades_tree(self):
        for row in self.grades_tree.get_children():
            self.grades_tree.delete(row)
        for s in self.students_data["students"]:
            for topic, grade in s.get("grades", {}).items():
                self.grades_tree.insert('', 'end', values=(s["name"], topic, grade))
    
    def refresh_debtors(self):
        for row in self.debtors_tree.get_children():
            self.debtors_tree.delete(row)
        for s in self.students_data["students"]:
            debts = [topic for topic, grade in s.get("grades", {}).items() if grade in ('2', '1', 'неуд', '0')]
            if debts:
                self.debtors_tree.insert('', 'end', values=(s["name"], ", ".join(debts)))
    
    def refresh_stats(self):
        self.stats_text.delete('1.0', 'end')
        all_grades = []
        student_avgs = []
        for s in self.students_data["students"]:
            grades = [int(g) for g in s.get("grades", {}).values() if g.isdigit()]
            if grades:
                avg = sum(grades) / len(grades)
                student_avgs.append((s["name"], avg))
                all_grades.extend(grades)
        
        overall = sum(all_grades) / len(all_grades) if all_grades else 0
        
        self.stats_text.insert('end', "📊 СТАТИСТИКА УСПЕВАЕМОСТИ\n\n")
        self.stats_text.insert('end', f"✅ Всего учеников: {len(self.students_data['students'])}\n")
        self.stats_text.insert('end', f"📈 Средний балл по предмету: {overall:.2f}\n\n")
        self.stats_text.insert('end', "Средний балл каждого ученика:\n")
        for name, avg in sorted(student_avgs, key=lambda x: x[1], reverse=True):
            self.stats_text.insert('end', f"   {name}: {avg:.2f}\n")
    
    def logout(self):
        self.current_user = None
        self.show_login()

# ==================== ЗАПУСК ====================
if __name__ == "__main__":
    root = tk.Tk()
    app = TeacherApp(root)
    root.mainloop()